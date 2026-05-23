#!/usr/bin/env python3
"""
2026 年度目標達成進度表產生器

讀取總合目標與區目標 Excel，產出四部（壯年部、婦人部、男子部、女子部）的 HTML 統計報表。

使用方式：
  1. 將最新的 Excel 檔案放入本腳本所在目錄
  2. 執行 python3 generate_report.py
  3. 產出的 HTML 報表在 output/ 目錄
"""

import glob
import os
import re
from datetime import datetime

import openpyxl
from jinja2 import Environment, FileSystemLoader

# ---------------------------------------------------------------------------
# 欄位對應：每個部別在 Excel 中的起始欄 (1-indexed)
# 每個目標類別佔 3 欄：目標, 達成, 比例
# ---------------------------------------------------------------------------
DEPARTMENTS = {
    '壯年部': {'個人資料卡': 3,  '御本尊': 18, 'My創新': 33, '實動會員數': 48},
    '婦人部': {'個人資料卡': 6,  '御本尊': 21, 'My創新': 36, '實動會員數': 51},
    '男子部': {'個人資料卡': 9,  '御本尊': 24, 'My創新': 39, '實動會員數': 54},
    '女子部': {'個人資料卡': 12, '御本尊': 27, 'My創新': 42, '實動會員數': 57},
}

CATEGORIES = ['個人資料卡', '御本尊', 'My創新', '實動會員數']

DATA_START_ROW = 6
DATA_END_ROW = 68  # 含總計列


# ---------------------------------------------------------------------------
# 工具函式
# ---------------------------------------------------------------------------

def rate_class(rate):
    """根據達成率回傳對應的 CSS class。"""
    if rate is None:
        return ''
    if rate >= 1.0:
        return 'rate-100'
    if rate >= 0.8:
        return 'rate-80'
    if rate >= 0.5:
        return 'rate-50'
    if rate >= 0.3:
        return 'rate-30'
    return ''


def safe_rate(numerator, denominator):
    """計算比率，分母為 0 時回傳 None。"""
    if denominator and denominator != 0:
        return numerator / denominator
    return None


def find_latest_files(directory):
    """找到目錄中日期最新的一組總合目標與區目標檔案。"""

    def extract_date(filepath):
        match = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(filepath))
        return datetime.strptime(match.group(1), '%Y-%m-%d') if match else datetime.min

    total_files = [f for f in glob.glob(os.path.join(directory, '*總合目標*.xlsx'))
                   if not os.path.basename(f).startswith('~$')]
    area_files = [f for f in glob.glob(os.path.join(directory, '*區目標*.xlsx'))
                  if not os.path.basename(f).startswith('~$')]

    if not total_files:
        raise FileNotFoundError("找不到總合目標 Excel 檔案")
    if not area_files:
        raise FileNotFoundError("找不到區目標 Excel 檔案")

    latest_total = max(total_files, key=extract_date)
    latest_area = max(area_files, key=extract_date)
    report_date = extract_date(latest_total).strftime('%Y-%m-%d')

    return latest_total, latest_area, report_date


def extract_department_data(ws_total, ws_area, col_map):
    """提取某一部的所有資料列。"""
    rows = []
    current_area = None
    group_idx = 0

    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        area = ws_total.cell(r, 1).value or ''
        district = ws_total.cell(r, 2).value or ''

        # 區域分組索引（用於交替底色）
        if area != current_area:
            if current_area is not None:
                group_idx += 1
            current_area = area

        row = {
            'area': area,
            'district': district,
            'is_total': (area == '總計'),
            'group_idx': group_idx,
            'categories': {},
        }

        for cat_name, start_col in col_map.items():
            base = ws_total.cell(r, start_col).value or 0       # 基本盤
            custom = ws_area.cell(r, start_col).value or 0       # 自訂目標
            done = ws_total.cell(r, start_col + 1).value or 0   # 達成
            rate = safe_rate(done, custom)                        # 達成率(vs 自訂)

            cat = {
                'base': base,
                'custom': custom,
                'done': done,
                'rate': rate,
            }

            # 個人資料卡：額外計算兩年三倍目標與達成率
            if cat_name == '個人資料卡':
                two_year_target = base * 2
                cat['two_year_target'] = two_year_target
                cat['two_year_rate'] = safe_rate(done, two_year_target)

            row['categories'][cat_name] = cat

        rows.append(row)

    return rows


def verify_data(ws_total, ws_area, dept_name, col_map):
    """雙重驗算：比對程式計算的達成率與 Excel 原始比例欄位。"""
    errors = []
    warnings = []

    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        district = ws_total.cell(r, 2).value or ws_total.cell(r, 1).value or f'Row{r}'

        for cat_name, start_col in col_map.items():
            # 驗算 1：兩份 Excel 的達成數應一致
            done_total = ws_total.cell(r, start_col + 1).value or 0
            done_area = ws_area.cell(r, start_col + 1).value or 0
            if done_total != done_area:
                warnings.append(
                    f"  [{district}] {cat_name} 達成數不一致: "
                    f"總合={done_total}, 區={done_area}"
                )

            # 驗算 2：程式算的達成率 vs Excel 原始比例欄 (start_col+2)
            goal = ws_total.cell(r, start_col).value or 0
            done = done_total
            excel_rate = ws_total.cell(r, start_col + 2).value  # Excel 的比例欄

            if goal and goal != 0 and excel_rate is not None:
                computed_rate = done / goal
                diff = abs(computed_rate - excel_rate)
                if diff > 0.005:  # 容許 0.5% 四捨五入誤差
                    errors.append(
                        f"  [{district}] {cat_name} 總合達成率偏差: "
                        f"程式={computed_rate:.4f}, Excel={excel_rate:.4f}"
                    )

            # 驗算 3：區目標的比例欄也交叉比對
            custom_goal = ws_area.cell(r, start_col).value or 0
            excel_area_rate = ws_area.cell(r, start_col + 2).value

            if custom_goal and custom_goal != 0 and excel_area_rate is not None:
                computed_area_rate = done_area / custom_goal
                diff = abs(computed_area_rate - excel_area_rate)
                if diff > 0.005:
                    errors.append(
                        f"  [{district}] {cat_name} 區達成率偏差: "
                        f"程式={computed_area_rate:.4f}, Excel={excel_area_rate:.4f}"
                    )

    # 輸出驗算結果
    print(f"\n  驗算結果 [{dept_name}]:")
    if warnings:
        print(f"    ⚠ 警告 ({len(warnings)} 筆):")
        for w in warnings[:10]:
            print(f"      {w}")
        if len(warnings) > 10:
            print(f"      ... 另有 {len(warnings) - 10} 筆")
    if errors:
        print(f"    ✗ 偏差 ({len(errors)} 筆):")
        for e in errors[:10]:
            print(f"      {e}")
        if len(errors) > 10:
            print(f"      ... 另有 {len(errors) - 10} 筆")
    if not warnings and not errors:
        print("    ✓ 全部通過")

    return errors, warnings


def compute_summary(rows):
    """計算儀表板所需的摘要數據。"""
    data_rows = [r for r in rows if not r['is_total'] and r['area'] != '圈以上']
    total_row = next((r for r in rows if r['is_total']), None)

    summary = {}
    for cat in CATEGORIES:
        total_done = total_row['categories'][cat]['done'] if total_row else 0
        total_custom = total_row['categories'][cat]['custom'] if total_row else 0
        total_rate = total_row['categories'][cat]['rate']

        # 排行榜：按達成率排序（排除無目標的）
        ranked = sorted(
            [
                {
                    'district': r['district'],
                    'area': r['area'],
                    'rate': r['categories'][cat]['rate'] or 0,
                    'done': r['categories'][cat]['done'],
                    'custom': r['categories'][cat]['custom'],
                }
                for r in data_rows
                if r['district'] and r['categories'][cat]['custom']
            ],
            key=lambda x: (-x['rate'], -x['done']),
        )

        # 達成率分佈
        all_rates = [(r['categories'][cat]['rate'] or 0) for r in data_rows if r['district']]
        dist = {
            'gte100': sum(1 for rt in all_rates if rt >= 1.0),
            'gte80': sum(1 for rt in all_rates if 0.8 <= rt < 1.0),
            'gte50': sum(1 for rt in all_rates if 0.5 <= rt < 0.8),
            'gte30': sum(1 for rt in all_rates if 0.3 <= rt < 0.5),
            'lt30': sum(1 for rt in all_rates if rt < 0.3),
        }

        cat_summary = {
            'total_done': total_done,
            'total_custom': total_custom,
            'total_rate': total_rate,
            'ranked': ranked,
            'distribution': dist,
            'total_areas': len(all_rates),
        }

        if cat == '個人資料卡' and total_row:
            cat_summary['two_year_target'] = total_row['categories'][cat].get('two_year_target', 0)
            cat_summary['two_year_rate'] = total_row['categories'][cat].get('two_year_rate')

        summary[cat] = cat_summary

    return summary


# ---------------------------------------------------------------------------
# 主程式
# ---------------------------------------------------------------------------

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # 1. 找到最新檔案
    total_file, area_file, report_date = find_latest_files(base_dir)
    print(f"總合目標: {os.path.basename(total_file)}")
    print(f"區目標:   {os.path.basename(area_file)}")
    print(f"報表日期: {report_date}")

    # 2. 讀取 Excel
    wb_total = openpyxl.load_workbook(total_file, data_only=True)
    wb_area = openpyxl.load_workbook(area_file, data_only=True)
    ws_total = wb_total.active
    ws_area = wb_area.active

    # 3. 載入模板
    env = Environment(loader=FileSystemLoader(base_dir))
    env.globals['rate_class'] = rate_class
    template = env.get_template('template.html')

    # 4. 產出報表
    docs_dir = os.path.join(base_dir, 'docs')
    history_dir = os.path.join(docs_dir, report_date)
    os.makedirs(docs_dir, exist_ok=True)
    os.makedirs(history_dir, exist_ok=True)

    for dept_name, col_map in DEPARTMENTS.items():
        # 雙重驗算
        verify_data(ws_total, ws_area, dept_name, col_map)

        rows = extract_department_data(ws_total, ws_area, col_map)
        summary = compute_summary(rows)

        html = template.render(
            dept_name=dept_name,
            report_date=report_date,
            rows=rows,
            categories=CATEGORIES,
            summary=summary,
        )

        filename = f'{dept_name}_統計.html'

        # 寫入最新版（docs/ 根目錄，供 GitHub Pages）
        with open(os.path.join(docs_dir, filename), 'w', encoding='utf-8') as f:
            f.write(html)

        # 寫入歷史版（docs/YYYY-MM-DD/）
        with open(os.path.join(history_dir, filename), 'w', encoding='utf-8') as f:
            f.write(html)

        print(f"  ✓ {filename}")

    print(f"\n最新報表: {docs_dir}/")
    print(f"歷史存檔: {history_dir}/")

    wb_total.close()
    wb_area.close()
    print("全部完成！")


if __name__ == '__main__':
    main()
